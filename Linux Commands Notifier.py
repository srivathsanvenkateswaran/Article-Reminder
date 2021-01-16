import notify2
import openpyxl

notify2.init("Linux Commands Notifier")
# Establishing a D Bus Connection

ICON_PATH ='/home/srivathsan/Coding Files/Cron Tabs Scripts/Linux Commands Notifier/linuxPenguinLogo.png'
# Path of the icon image

n = notify2.Notification(None, icon=ICON_PATH)
# Creating a notification object and passing the icon as as attribute 
# NOTE: We set the Title of the Notification as None because we will update it later [Need to fetch it from a Spreadsheet]

n.set_urgency(notify2.URGENCY_NORMAL)
# Setting the urgency level of the notification

n.set_timeout(1000)
# Setting the timeout for notification 

wb = openpyxl.load_workbook('Linux Commands.xlsx')
# Opening the Excel file 

sheet = wb.active
# Getting the Active sheet 

TITLE_COLUMN = 1
DESCRIPTION_COLUMN = 2
# We have decided the column variable as constants

CURRENT_ROW = 1
# We create an integer variable CURRENT_ROW

Title = sheet.cell(row = CURRENT_ROW, column = TITLE_COLUMN).value
Description = sheet.cell(row = CURRENT_ROW, column = DESCRIPTION_COLUMN).value
# We get the value of Title and Description using the appropriate values of row and column

sheet.delete_rows(CURRENT_ROW, amount = 1)
# Delete the first row

wb.save('Linux Commands.xlsx')
# Save the Sheet 

n.update(Title, Description)
#Updating the Title and Description into the notification object

n.show()
#This will show the notification 